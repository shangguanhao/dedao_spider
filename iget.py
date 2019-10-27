import openpyxl
from appium import webdriver
import time

from selenium.webdriver.support.wait import WebDriverWait

from config import DRIVER_SERVER, TIMEOUT, FLICK_START_X, FLICK_DISTANCE, FLICK_START_Y, PATH


class CourseData(object):
    def __init__(self, course_name, summary, lecturer_name_and_title, price, subscribe_num):
        super().__init__()
        self.course_name = course_name
        self.summary = summary
        self.lecturer_name_and_title = lecturer_name_and_title
        self.price = price
        self.subscribe_num = subscribe_num

    def __eq__(self, other):
        if self.course_name != other.course_name:
            return False
        if self.summary != other.summary:
            return False
        if self.lecturer_name_and_title != other.lecturer_name_and_title:
            return False
        if self.price != other.price:
            return False
        if self.subscribe_num != other.subscribe_num:
            return False
        return True


class Action():
    def __init__(self):
        print("init")
        self.desired_caps = {
            "platformName": "Android",
            "deviceName": "127.0.0.1:62001",
            'platformVersion': '5.1',
            "appPackage": "com.luojilab.player",
            "appActivity": "com.luojilab.business.HomeTabActivity",  # 主页
            "noReset": True
        }
        self.driver = webdriver.Remote(DRIVER_SERVER, self.desired_caps)
        self.wait = WebDriverWait(self.driver, TIMEOUT)

    def entry(self):
        college_list = ["商学院", "能力学院", "视野学院", "人文社科", "科学学院"]
        wait = WebDriverWait(self.driver, 20)

        # 同意隐私条款
        print("agree")
        btn_aggree = self.driver.find_elements_by_id('com.luojilab.player:id/btn_agree')
        if btn_aggree != []:
            btn_aggree[0].click()

        # 定义数据输出的Excel
        try:
            wb = openpyxl.load_workbook(PATH)
        except:
            wb = openpyxl.Workbook()
        # wb = openpyxl.Workbook()
        sheet_name = time.strftime("%F")
        sheet = wb.create_sheet(sheet_name)
        row = ["学院", "课程名称", "课程摘要", "主讲人", "单价", "销量", "销售金额"]
        sheet.append(row)
        sheet.column_dimensions['A'].width = 15  # 学院
        sheet.column_dimensions['B'].width = 40  # 标题
        sheet.column_dimensions['C'].width = 40  # 摘要
        sheet.column_dimensions['D'].width = 40  # 作者
        sheet.column_dimensions['E'].width = 15  # 单价
        sheet.column_dimensions['F'].width = 15  # 销量
        sheet.column_dimensions['G'].width = 15  # 销售金额

        for col in college_list:
            print("enter course...")
            time.sleep(1)
            course = self.driver.find_element_by_xpath("//android.widget.TextView[@text='课程']")  # 课程
            course.click()

            print("get college...")
            time.sleep(1)
            colleges = self.driver.find_element_by_id(
                'com.luojilab.player:id/college_filter_list').find_elements_by_xpath(
                '//android.widget.TextView')
            for college in colleges:
                print(college.text)
                if col in college.text:
                    print(col, "click")
                    shang_course_names = []
                    shang_summaries = []
                    shang_lecturer_name_and_titles = []
                    shang_prices = []
                    shang_subscribe_nums = []
                    college.click()
                    while True:
                        time.sleep(2)
                        temps = self.driver.find_element_by_id('com.luojilab.player:id/rv').find_elements_by_class_name(
                            'android.widget.RelativeLayout')
                        for temp in temps:
                            try:
                                course_names = temp.find_element_by_id('com.luojilab.player:id/column_name')
                                summaries = temp.find_element_by_id('com.luojilab.player:id/summary')
                                lecturer_name_and_titles = temp.find_element_by_id(
                                    'com.luojilab.player:id/tv_name_and_title')
                                prices = temp.find_element_by_id('com.luojilab.player:id/price')
                                subscribe_nums = temp.find_element_by_id('com.luojilab.player:id/tv_subscribe_num')
                            except Exception as e:
                                continue

                            shang_course_names.append(course_names.text)
                            shang_summaries.append(summaries.text)
                            shang_lecturer_name_and_titles.append(lecturer_name_and_titles.text)
                            shang_prices.append(prices.text)
                            shang_subscribe_nums.append(subscribe_nums.text)

                        if temps[-1].find_elements_by_id('com.luojilab.player:id/rx_loadmore_text') != []:
                            break
                        else:
                            self.driver.swipe(FLICK_START_X, FLICK_START_Y, FLICK_START_X,
                                              FLICK_START_Y + FLICK_DISTANCE,
                                              2000)

                    print(len(shang_course_names), shang_course_names)
                    print(len(shang_summaries), shang_summaries)
                    print(len(shang_lecturer_name_and_titles), shang_lecturer_name_and_titles)
                    print(len(shang_prices), shang_prices)
                    print(len(shang_subscribe_nums), shang_prices)

                    course_data_list = []
                    if len(shang_course_names) == len(shang_summaries) and len(shang_course_names) == len(
                            shang_lecturer_name_and_titles) and len(shang_course_names) == len(shang_prices) and len(
                        shang_course_names) == len(shang_subscribe_nums):
                        for i in range(len(shang_course_names)):
                            course_data = CourseData(shang_course_names[i], shang_summaries[i],
                                                     shang_lecturer_name_and_titles[i], shang_prices[i],
                                                     shang_subscribe_nums[i])
                            if course_data not in course_data_list:
                                course_data_list.append(course_data)

                    print(len(course_data_list))

                    for i in range(len(course_data_list)):
                        print(course_data_list[i].course_name, "-", course_data_list[i].summary, "-",
                              course_data_list[i].lecturer_name_and_title, "-", course_data_list[i].price, "-",
                              course_data_list[i].subscribe_num)
                        time.sleep(0.1)

                        # 写入Excel
                        cell_price = course_data_list[i].price[course_data_list[i].price.strip().find("¥") + 2:]
                        cell_subscribe_num = int(
                            course_data_list[i].subscribe_num[:course_data_list[i].subscribe_num.strip().find("人")])
                        row = [col, course_data_list[i].course_name, course_data_list[i].summary,
                               course_data_list[i].lecturer_name_and_title, cell_price,
                               cell_subscribe_num, float(cell_price) * cell_subscribe_num]
                        sheet.append(row)
                        wb.save(PATH)
                    break
            btn_back = self.driver.find_elements_by_id('com.luojilab.player:id/iv_back_btn')
            if btn_back != []:
                btn_back[0].click()


if __name__ == '__main__':
    action = Action()
    action.entry()
